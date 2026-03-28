"""
Script para criar planilha base de controle de estoque para salão de beleza.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from datetime import date

wb = openpyxl.Workbook()

# ── Estilos ──────────────────────────────────────────────────────────────────
ROSA = "D4638F"
ROSA_CLARO = "F5E1EB"
BRANCO = "FFFFFF"
CINZA = "F2F2F2"
VERMELHO = "FF4444"
AMARELO = "FFD700"
VERDE = "4CAF50"

header_font = Font(name="Calibri", bold=True, color=BRANCO, size=11)
header_fill = PatternFill(start_color=ROSA, end_color=ROSA, fill_type="solid")
title_font = Font(name="Calibri", bold=True, color=ROSA, size=16)
subtitle_font = Font(name="Calibri", bold=True, color=ROSA, size=12)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
alt_fill = PatternFill(start_color=ROSA_CLARO, end_color=ROSA_CLARO, fill_type="solid")
alerta_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border


def style_data_area(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = center
            if (r - start_row) % 2 == 1:
                cell.fill = PatternFill(start_color=CINZA, end_color=CINZA, fill_type="solid")


def auto_width(ws, max_col, min_width=14):
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = min_width


# ═══════════════════════════════════════════════════════════════════════════════
# ABA 1 — ESTOQUE DE PRODUTOS
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Estoque de Produtos"
ws1.sheet_properties.tabColor = ROSA

headers1 = [
    "Código", "Produto", "Categoria", "Marca", "Unidade",
    "Qtd. Atual", "Qtd. Mínima", "Status", "Preço Custo (R$)",
    "Preço Venda (R$)", "Fornecedor", "Validade", "Localização",
]

# Título
ws1.merge_cells("A1:M1")
ws1.cell(row=1, column=1).value = "CONTROLE DE ESTOQUE — SALÃO DE BELEZA"
ws1.cell(row=1, column=1).font = title_font
ws1.cell(row=1, column=1).alignment = center
ws1.row_dimensions[1].height = 40

# Cabeçalhos
for col, h in enumerate(headers1, 1):
    ws1.cell(row=3, column=col).value = h
style_header(ws1, 3, len(headers1))

# Dados de exemplo
categorias = {
    "Cabelo": [
        ("P001", "Shampoo Profissional 1L", "Cabelo", "L'Oréal", "un", 15, 5, "", 32.00, 65.00, "Dist. Beauty", "2027-03-01", "Prateleira A1"),
        ("P002", "Condicionador Profissional 1L", "Cabelo", "L'Oréal", "un", 12, 5, "", 35.00, 70.00, "Dist. Beauty", "2027-03-01", "Prateleira A1"),
        ("P003", "Máscara de Hidratação 500g", "Cabelo", "Kerastase", "un", 8, 3, "", 55.00, 120.00, "Dist. Beauty", "2027-06-01", "Prateleira A2"),
        ("P004", "Tintura Creme (diversas cores)", "Cabelo", "Wella", "un", 40, 15, "", 18.00, 45.00, "Color Supply", "2027-12-01", "Armário B1"),
        ("P005", "Água Oxigenada 20 vol. 1L", "Cabelo", "Wella", "un", 10, 5, "", 12.00, 28.00, "Color Supply", "2027-09-01", "Armário B1"),
        ("P006", "Pó Descolorante 500g", "Cabelo", "Alfaparf", "un", 6, 3, "", 40.00, 85.00, "Color Supply", "2027-08-01", "Armário B2"),
        ("P007", "Óleo Reparador 100ml", "Cabelo", "Moroccanoil", "un", 10, 4, "", 65.00, 140.00, "Dist. Beauty", "2028-01-01", "Prateleira A2"),
        ("P008", "Spray Fixador 400ml", "Cabelo", "Kérastase", "un", 7, 3, "", 42.00, 90.00, "Dist. Beauty", "2027-10-01", "Prateleira A3"),
    ],
    "Unha": [
        ("P009", "Esmalte (diversas cores)", "Unha", "Risqué", "un", 50, 20, "", 4.50, 12.00, "Beleza Total", "2028-06-01", "Gaveta C1"),
        ("P010", "Base Fortalecedora", "Unha", "Risqué", "un", 10, 5, "", 5.00, 14.00, "Beleza Total", "2028-06-01", "Gaveta C1"),
        ("P011", "Removedor de Esmalte 500ml", "Unha", "Farmax", "un", 8, 3, "", 8.00, 18.00, "Beleza Total", "2028-01-01", "Gaveta C2"),
        ("P012", "Acetona 1L", "Unha", "Farmax", "un", 5, 2, "", 10.00, 22.00, "Beleza Total", "2028-03-01", "Gaveta C2"),
    ],
    "Pele/Estética": [
        ("P013", "Creme Hidratante Corporal 1L", "Pele/Estética", "Nivea", "un", 6, 3, "", 25.00, 55.00, "Dist. Beauty", "2027-11-01", "Prateleira D1"),
        ("P014", "Protetor Solar FPS 50 200ml", "Pele/Estética", "La Roche", "un", 5, 2, "", 45.00, 95.00, "Dist. Beauty", "2027-07-01", "Prateleira D1"),
        ("P015", "Cera Depilatória 1kg", "Pele/Estética", "Depil Bella", "un", 4, 2, "", 30.00, 65.00, "Estética Plus", "2027-12-01", "Armário E1"),
    ],
    "Descartáveis": [
        ("P016", "Luvas Descartáveis (cx 100)", "Descartáveis", "Descarpack", "cx", 8, 3, "", 28.00, 0, "Higiene Pro", "2028-12-01", "Armário F1"),
        ("P017", "Touca Descartável (pct 100)", "Descartáveis", "Descarpack", "pct", 5, 2, "", 15.00, 0, "Higiene Pro", "2028-12-01", "Armário F1"),
        ("P018", "Papel Alumínio (rolo 100m)", "Descartáveis", "Wyda", "rolo", 6, 3, "", 22.00, 0, "Higiene Pro", "—", "Armário F2"),
        ("P019", "Algodão 500g", "Descartáveis", "Apolo", "pct", 10, 4, "", 12.00, 0, "Higiene Pro", "—", "Armário F2"),
        ("P020", "Lixa de Unha (pct 50)", "Descartáveis", "Marco Boni", "pct", 4, 2, "", 18.00, 0, "Beleza Total", "—", "Gaveta C3"),
    ],
}

row = 4
for cat, items in categorias.items():
    for item in items:
        for col, val in enumerate(item, 1):
            ws1.cell(row=row, column=col).value = val
        # Fórmula de status
        ws1.cell(row=row, column=8).value = f'=IF(F{row}<=G{row},"⚠ REPOR",IF(F{row}<=G{row}*1.5,"ATENÇÃO","OK"))'
        # Formato moeda
        ws1.cell(row=row, column=9).number_format = '#,##0.00'
        ws1.cell(row=row, column=10).number_format = '#,##0.00'
        row += 1

last_data_row = row - 1
style_data_area(ws1, 4, last_data_row, len(headers1))

# Formatação condicional para status
ws1.conditional_formatting.add(
    f"H4:H{last_data_row}",
    CellIsRule(operator="equal", formula=['"⚠ REPOR"'], fill=PatternFill(bgColor="FF9999"))
)
ws1.conditional_formatting.add(
    f"H4:H{last_data_row}",
    CellIsRule(operator="equal", formula=['"ATENÇÃO"'], fill=PatternFill(bgColor="FFFF99"))
)
ws1.conditional_formatting.add(
    f"H4:H{last_data_row}",
    CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill(bgColor="99FF99"))
)

# Resumo
summary_row = last_data_row + 2
ws1.cell(row=summary_row, column=1).value = "RESUMO"
ws1.cell(row=summary_row, column=1).font = subtitle_font
ws1.cell(row=summary_row + 1, column=1).value = "Total de Itens Cadastrados:"
ws1.cell(row=summary_row + 1, column=2).value = f"=COUNTA(A4:A{last_data_row})"
ws1.cell(row=summary_row + 2, column=1).value = "Itens para Repor:"
ws1.cell(row=summary_row + 2, column=2).value = f'=COUNTIF(H4:H{last_data_row},"⚠ REPOR")'
ws1.cell(row=summary_row + 3, column=1).value = "Valor Total em Estoque (Custo):"
ws1.cell(row=summary_row + 3, column=2).value = f"=SUMPRODUCT(F4:F{last_data_row},I4:I{last_data_row})"
ws1.cell(row=summary_row + 3, column=2).number_format = 'R$ #,##0.00'

auto_width(ws1, len(headers1), 15)
ws1.column_dimensions["B"].width = 32
ws1.column_dimensions["K"].width = 18

# Filtro automático
ws1.auto_filter.ref = f"A3:M{last_data_row}"

# Congelar painel
ws1.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════════════════════════
# ABA 2 — MOVIMENTAÇÕES (ENTRADAS E SAÍDAS)
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Movimentações")
ws2.sheet_properties.tabColor = "8E44AD"

ws2.merge_cells("A1:I1")
ws2.cell(row=1, column=1).value = "REGISTRO DE MOVIMENTAÇÕES"
ws2.cell(row=1, column=1).font = title_font
ws2.cell(row=1, column=1).alignment = center
ws2.row_dimensions[1].height = 40

headers2 = [
    "Data", "Código Produto", "Produto", "Tipo (Entrada/Saída)",
    "Quantidade", "Motivo", "Responsável", "Nº Nota Fiscal", "Observações",
]
for col, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=col).value = h
style_header(ws2, 3, len(headers2))

# Exemplos
movs = [
    ("2026-03-25", "P001", "Shampoo Profissional 1L", "Entrada", 10, "Compra", "Maria", "NF-4521", ""),
    ("2026-03-25", "P004", "Tintura Creme", "Entrada", 20, "Compra", "Maria", "NF-4522", ""),
    ("2026-03-26", "P001", "Shampoo Profissional 1L", "Saída", 2, "Uso interno", "Ana", "", "Cliente VIP"),
    ("2026-03-26", "P009", "Esmalte", "Saída", 3, "Uso interno", "Juliana", "", ""),
    ("2026-03-27", "P007", "Óleo Reparador 100ml", "Saída", 1, "Venda direta", "Ana", "", ""),
]
for i, mov in enumerate(movs):
    for col, val in enumerate(mov, 1):
        ws2.cell(row=4 + i, column=col).value = val
style_data_area(ws2, 4, 8 + 50, len(headers2))

# Formatação condicional: Entrada = verde, Saída = rosa
ws2.conditional_formatting.add(
    "D4:D200",
    CellIsRule(operator="equal", formula=['"Entrada"'], fill=PatternFill(bgColor="C6EFCE"))
)
ws2.conditional_formatting.add(
    "D4:D200",
    CellIsRule(operator="equal", formula=['"Saída"'], fill=PatternFill(bgColor="FFC7CE"))
)

auto_width(ws2, len(headers2), 16)
ws2.column_dimensions["C"].width = 30
ws2.column_dimensions["F"].width = 20
ws2.auto_filter.ref = "A3:I3"
ws2.freeze_panes = "A4"

# Validação de dados para Tipo
from openpyxl.worksheet.datavalidation import DataValidation
dv_tipo = DataValidation(type="list", formula1='"Entrada,Saída"', allow_blank=True)
dv_tipo.prompt = "Selecione o tipo"
dv_tipo.promptTitle = "Tipo de Movimentação"
ws2.add_data_validation(dv_tipo)
dv_tipo.add("D4:D500")

dv_motivo = DataValidation(type="list", formula1='"Compra,Uso interno,Venda direta,Devolução,Perda/Avaria,Amostra"', allow_blank=True)
ws2.add_data_validation(dv_motivo)
dv_motivo.add("F4:F500")


# ═══════════════════════════════════════════════════════════════════════════════
# ABA 3 — FORNECEDORES
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Fornecedores")
ws3.sheet_properties.tabColor = "2ECC71"

ws3.merge_cells("A1:H1")
ws3.cell(row=1, column=1).value = "CADASTRO DE FORNECEDORES"
ws3.cell(row=1, column=1).font = title_font
ws3.cell(row=1, column=1).alignment = center
ws3.row_dimensions[1].height = 40

headers3 = [
    "Código", "Fornecedor", "CNPJ", "Contato", "Telefone",
    "E-mail", "Categorias", "Prazo Entrega (dias)",
]
for col, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=col).value = h
style_header(ws3, 3, len(headers3))

fornecedores = [
    ("F001", "Dist. Beauty", "12.345.678/0001-99", "Carlos", "(11) 99999-1111", "contato@distbeauty.com", "Cabelo, Pele", 5),
    ("F002", "Color Supply", "23.456.789/0001-88", "Fernanda", "(11) 98888-2222", "vendas@colorsupply.com", "Cabelo (Coloração)", 3),
    ("F003", "Beleza Total", "34.567.890/0001-77", "Roberto", "(11) 97777-3333", "pedidos@belezatotal.com", "Unha, Descartáveis", 7),
    ("F004", "Estética Plus", "45.678.901/0001-66", "Luciana", "(11) 96666-4444", "comercial@esteticaplus.com", "Pele/Estética", 4),
    ("F005", "Higiene Pro", "56.789.012/0001-55", "André", "(11) 95555-5555", "vendas@higienepro.com", "Descartáveis", 2),
]
for i, f in enumerate(fornecedores):
    for col, val in enumerate(f, 1):
        ws3.cell(row=4 + i, column=col).value = val
style_data_area(ws3, 4, 8 + 20, len(headers3))
auto_width(ws3, len(headers3), 18)
ws3.column_dimensions["B"].width = 22
ws3.column_dimensions["F"].width = 30
ws3.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════════════════════════
# ABA 4 — LISTA DE COMPRAS
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Lista de Compras")
ws4.sheet_properties.tabColor = "E67E22"

ws4.merge_cells("A1:H1")
ws4.cell(row=1, column=1).value = "LISTA DE COMPRAS PENDENTES"
ws4.cell(row=1, column=1).font = title_font
ws4.cell(row=1, column=1).alignment = center
ws4.row_dimensions[1].height = 40

headers4 = [
    "Data Pedido", "Código Produto", "Produto", "Qtd. a Comprar",
    "Fornecedor", "Preço Estimado (R$)", "Status", "Data Recebimento",
]
for col, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=col).value = h
style_header(ws4, 3, len(headers4))
style_data_area(ws4, 4, 4 + 50, len(headers4))
auto_width(ws4, len(headers4), 18)
ws4.column_dimensions["C"].width = 30

dv_status_compra = DataValidation(type="list", formula1='"Pendente,Pedido Feito,Recebido,Cancelado"', allow_blank=True)
ws4.add_data_validation(dv_status_compra)
dv_status_compra.add("G4:G500")

ws4.conditional_formatting.add(
    "G4:G200",
    CellIsRule(operator="equal", formula=['"Pendente"'], fill=PatternFill(bgColor="FFFF99"))
)
ws4.conditional_formatting.add(
    "G4:G200",
    CellIsRule(operator="equal", formula=['"Recebido"'], fill=PatternFill(bgColor="99FF99"))
)
ws4.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════════════════════════
# ABA 5 — DASHBOARD / RESUMO
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Dashboard")
ws5.sheet_properties.tabColor = "3498DB"

ws5.merge_cells("A1:F1")
ws5.cell(row=1, column=1).value = "DASHBOARD — VISÃO GERAL DO ESTOQUE"
ws5.cell(row=1, column=1).font = title_font
ws5.cell(row=1, column=1).alignment = center
ws5.row_dimensions[1].height = 40

# KPIs
kpis = [
    ("Total de Produtos Cadastrados", f"=COUNTA('Estoque de Produtos'!A4:A{last_data_row})"),
    ("Produtos com Estoque OK", f"=COUNTIF('Estoque de Produtos'!H4:H{last_data_row},\"OK\")"),
    ("Produtos em ATENÇÃO", f"=COUNTIF('Estoque de Produtos'!H4:H{last_data_row},\"ATENÇÃO\")"),
    ("Produtos para REPOR", f"=COUNTIF('Estoque de Produtos'!H4:H{last_data_row},\"⚠ REPOR\")"),
    ("Valor Total Estoque (Custo)", f"=SUMPRODUCT('Estoque de Produtos'!F4:F{last_data_row},'Estoque de Produtos'!I4:I{last_data_row})"),
    ("Valor Total Estoque (Venda)", f"=SUMPRODUCT('Estoque de Produtos'!F4:F{last_data_row},'Estoque de Produtos'!J4:J{last_data_row})"),
    ("Fornecedores Cadastrados", "=COUNTA(Fornecedores!A4:A100)"),
    ("Entradas no Período", "=COUNTIF(Movimentações!D4:D500,\"Entrada\")"),
    ("Saídas no Período", "=COUNTIF(Movimentações!D4:D500,\"Saída\")"),
]

for i, (label, formula) in enumerate(kpis):
    r = 3 + i
    ws5.cell(row=r, column=1).value = label
    ws5.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11)
    ws5.cell(row=r, column=1).alignment = left
    ws5.cell(row=r, column=2).value = formula
    ws5.cell(row=r, column=2).alignment = center
    ws5.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=14, color=ROSA)
    ws5.cell(row=r, column=2).border = thin_border
    if "Valor" in label:
        ws5.cell(row=r, column=2).number_format = 'R$ #,##0.00'

ws5.column_dimensions["A"].width = 35
ws5.column_dimensions["B"].width = 22

# Resumo por categoria
cat_row = 14
ws5.cell(row=cat_row, column=1).value = "RESUMO POR CATEGORIA"
ws5.cell(row=cat_row, column=1).font = subtitle_font
headers_cat = ["Categoria", "Qtd. Produtos", "Valor Estoque (Custo)"]
for col, h in enumerate(headers_cat, 1):
    ws5.cell(row=cat_row + 1, column=col).value = h
style_header(ws5, cat_row + 1, 3)

cats = ["Cabelo", "Unha", "Pele/Estética", "Descartáveis"]
for i, cat in enumerate(cats):
    r = cat_row + 2 + i
    ws5.cell(row=r, column=1).value = cat
    ws5.cell(row=r, column=1).border = thin_border
    ws5.cell(row=r, column=2).value = f"=COUNTIF('Estoque de Produtos'!C4:C{last_data_row},\"{cat}\")"
    ws5.cell(row=r, column=2).border = thin_border
    ws5.cell(row=r, column=2).alignment = center
    ws5.cell(row=r, column=3).value = f"=SUMPRODUCT(('Estoque de Produtos'!C4:C{last_data_row}=\"{cat}\")*'Estoque de Produtos'!F4:F{last_data_row}*'Estoque de Produtos'!I4:I{last_data_row})"
    ws5.cell(row=r, column=3).border = thin_border
    ws5.cell(row=r, column=3).number_format = 'R$ #,##0.00'
    ws5.cell(row=r, column=3).alignment = center

ws5.column_dimensions["C"].width = 25


# ═══════════════════════════════════════════════════════════════════════════════
# SALVAR
# ═══════════════════════════════════════════════════════════════════════════════
filename = "Estoque_Salao_de_Beleza.xlsx"
wb.save(filename)
print(f"Planilha '{filename}' criada com sucesso!")
